"""
Generates the blank Financials_Template.xlsx that unlisted-company users
fill in and upload. Ported unchanged from the nested create_template()
function in the old app's Unlisted Company Streamlit branch — pure
openpyxl generation, no UI dependency, so it moved over as-is.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BS_ITEMS = [
    'Equity and Liabilities', 'Equity', 'Share Capital', 'Reserves and Surplus', 'Other Equity',
    'Total Equity', 'Liabilities', 'Non-current Liabilities', 'Long Term Borrowings',
    'Net Deferred Tax Liabilities', 'Other Long Term Liabilities', 'Long Term Provisions',
    'Total Non-current Liabilities', 'Current Liabilities', 'Short Term Borrowings',
    'Trade Payables', 'Other Current Liabilities', 'Short Term Provisions',
    'Total Current Liabilities', 'Total Equity and Liabilities', '',
    'Assets', 'Non-current Assets', 'Tangible Assets', 'Intangible Assets',
    'Capital Work in Progress', 'Non-current Investments', 'Long Term Loans and Advances',
    'Other Non-current Assets', 'Total Non-current Assets', 'Current Assets',
    'Inventories', 'Trade Receivables', 'Cash and Bank Balances',
    'Short Term Loans and Advances', 'Other Current Assets', 'Total Current Assets', 'Total Assets'
]

BS_BOLD = {
    'Equity and Liabilities', 'Equity', 'Liabilities', 'Non-current Liabilities',
    'Current Liabilities', 'Assets', 'Non-current Assets', 'Current Assets',
}

PL_ITEMS = [
    'Net Revenue', 'Revenue Growth', 'Operating Cost', 'Cost of Materials Consumed',
    'Purchases of Stock-in-trade', 'Changes in Inventories / Finished Goods',
    'Employee Benefit Expense', 'Other Expenses', 'Total Operating Cost', 'EBITDA',
    'Other Income', 'Depreciation and Amortization Expense', 'Profit Before Interest and Tax',
    'Finance Costs', 'Profit Before Tax and Exceptional Items Before Tax',
    'Exceptional Items Before Tax', 'Profit Before Tax', 'Income Tax',
    'Profit for the Period from Continuing Operations'
]

PL_BOLD = {
    'Net Revenue', 'Total Operating Cost', 'EBITDA', 'Profit Before Interest and Tax',
    'Profit Before Tax', 'Profit for the Period from Continuing Operations',
}

HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _style_header_row(ws):
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = Font(bold=True, color='FFFFFF')
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(horizontal='center')


def create_unlisted_template(years=(23, 24, 25)):
    """Returns the .xlsx file as raw bytes, ready to send as a download."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    ws_bs = wb.create_sheet('BalanceSheet')
    ws_bs['A1'] = 'BALANCE SHEET'
    ws_bs['B1'], ws_bs['C1'], ws_bs['D1'] = years
    _style_header_row(ws_bs)

    row = 2
    for item in BS_ITEMS:
        ws_bs[f'A{row}'] = item
        if item in BS_BOLD:
            ws_bs[f'A{row}'].font = Font(bold=True)
        row += 1

    ws_bs.column_dimensions['A'].width = 35
    for col in ('B', 'C', 'D'):
        ws_bs.column_dimensions[col].width = 15

    ws_pl = wb.create_sheet('Profit&Loss')
    ws_pl['A1'] = 'PROFIT & LOSS'
    ws_pl['B1'], ws_pl['C1'], ws_pl['D1'] = years
    _style_header_row(ws_pl)

    row = 2
    for item in PL_ITEMS:
        ws_pl[f'A{row}'] = item
        if item in PL_BOLD:
            ws_pl[f'A{row}'].font = Font(bold=True)
        row += 1

    ws_pl.column_dimensions['A'].width = 50
    for col in ('B', 'C', 'D'):
        ws_pl.column_dimensions[col].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
