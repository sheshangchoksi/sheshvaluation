"""
Screener Excel Mode logic — ported from screener_excel_mode.py, which
is the module the ORIGINAL app's UI actually imported and used (an
earlier port of this file mistakenly used screener_excel_handler.py, a
simpler/incomplete alternate module that was never wired into the real
app — that was a real bug, not a simplification; this replaces it).

Covers: parsing Screener.in-format Excel (BalanceSheet/P&L sheets),
extracting financials AND shares-outstanding/face-value directly from
the sheet, optional Yahoo Finance lookup for current price + beta, and
Screener-specific DDM/RIM valuation math.
"""
import numpy as np
import pandas as pd

from app.logic._st_shim import st
from app.logic.yf_ratelimit import safe_ticker


"""
SCREENER EXCEL MODE - Auxiliary Module for SheshUltimate DCF Valuation System
==============================================================================

This module handles Screener.in template format Excel files with:
- Balance Sheet (rows with item names in column A, years in row 2)
- Profit and Loss Account (rows with item names in column A, years in row 2)

Features:
- Full DCF Valuation
- DDM (Dividend Discount Model) 
- RIM (Residual Income Model)
- Comparative Valuation
- Excel Download Functionality
- Ticker Input for Yahoo Finance (current price + beta)
- Current price vs fair values comparison
- Charts and visualizations from Listed mode

Author: SheshUltimate Team
Version: 2.0.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime


# ================================
# SCREENER EXCEL PARSING FUNCTIONS
# ================================

def parse_screener_excel_to_dataframes(excel_file):
    """
    Parse Screener.in template Excel file
    
    Expected Structure:
    - Sheet 1: 'Balance Sheet' with years in Row 2, items in Column A
    - Sheet 2: 'Profit and Loss Account' with years in Row 2, items in Column A
    
    Returns:
        df_bs, df_pl: DataFrames with Item column and year columns
    """
    try:
        # Read both sheets - note exact sheet names from template
        df_bs = pd.read_excel(excel_file, sheet_name='Balance Sheet', header=None)
        df_pl = pd.read_excel(excel_file, sheet_name='Profit and Loss Account', header=None)
        
        # Process Balance Sheet
        # Row 0: Title "BALANCE SHEET"
        # Row 1: "Report Date" + actual dates
        # Row 2+: Item names in column 0, values in columns 1+
        
        # Get years from Row 1 (index 1)
        bs_dates = df_bs.iloc[1, 1:].values  # Skip column 0 ("Report Date")
        
        # Convert dates to year strings
        bs_years = []
        for date_val in bs_dates:
            if pd.notna(date_val):
                if isinstance(date_val, datetime):
                    bs_years.append(f'_{date_val.year}')
                elif isinstance(date_val, str):
                    # Try to extract year from string
                    try:
                        year = pd.to_datetime(date_val).year
                        bs_years.append(f'_{year}')
                    except:
                        bs_years.append(f'_col{len(bs_years)+1}')
                else:
                    bs_years.append(f'_col{len(bs_years)+1}')
            else:
                bs_years.append(f'_col{len(bs_years)+1}')
        
        # Same for P&L
        pl_dates = df_pl.iloc[1, 1:].values
        pl_years = []
        for date_val in pl_dates:
            if pd.notna(date_val):
                if isinstance(date_val, datetime):
                    pl_years.append(f'_{date_val.year}')
                elif isinstance(date_val, str):
                    try:
                        year = pd.to_datetime(date_val).year
                        pl_years.append(f'_{year}')
                    except:
                        pl_years.append(f'_col{len(pl_years)+1}')
                else:
                    pl_years.append(f'_col{len(pl_years)+1}')
            else:
                pl_years.append(f'_col{len(pl_years)+1}')
        
        # Create column names: 'Item' + year columns
        bs_columns = ['Item'] + bs_years
        pl_columns = ['Item'] + pl_years
        
        # Extract data starting from Row 2 (index 2)
        df_bs_data = df_bs.iloc[2:].copy()
        df_pl_data = df_pl.iloc[2:].copy()
        
        # Limit to actual columns
        df_bs_data = df_bs_data.iloc[:, :len(bs_columns)]
        df_pl_data = df_pl_data.iloc[:, :len(pl_columns)]
        
        # Set column names
        df_bs_data.columns = bs_columns
        df_pl_data.columns = pl_columns
        
        # Reset index
        df_bs_data = df_bs_data.reset_index(drop=True)
        df_pl_data = df_pl_data.reset_index(drop=True)
        
        # Convert year columns to numeric
        for col in bs_years:
            df_bs_data[col] = pd.to_numeric(df_bs_data[col], errors='coerce').fillna(0.0)
        
        for col in pl_years:
            df_pl_data[col] = pd.to_numeric(df_pl_data[col], errors='coerce').fillna(0.0)
        
        # Remove rows where Item is NaN or empty
        df_bs_data = df_bs_data[df_bs_data['Item'].notna() & (df_bs_data['Item'] != '')]
        df_pl_data = df_pl_data[df_pl_data['Item'].notna() & (df_pl_data['Item'] != '')]
        
        # Clean Item column - remove whitespace
        df_bs_data['Item'] = df_bs_data['Item'].astype(str).str.strip()
        df_pl_data['Item'] = df_pl_data['Item'].astype(str).str.strip()
        
        return df_bs_data, df_pl_data
        
    except Exception as e:
        st.error(f"❌ Error parsing Screener Excel: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, None


def get_value_from_screener_df(df, item_name, year_col):
    """
    Extract value from Screener DataFrame by item name (case-insensitive match)
    Tries exact match first, then partial match
    
    Args:
        df: DataFrame with 'Item' column and year columns
        item_name: Item to search for
        year_col: Year column name (e.g., '_2023')
    
    Returns:
        float: Value or 0.0 if not found
    """
    if df is None or df.empty:
        return 0.0
    
    item_name_lower = item_name.lower()
    
    # Try exact match first (case-insensitive)
    exact_mask = df['Item'].str.lower().str.strip() == item_name_lower
    exact_matching = df[exact_mask]
    
    if not exact_matching.empty and year_col in exact_matching.columns:
        return float(exact_matching.iloc[0][year_col])
    
    # Fall back to partial match (contains)
    mask = df['Item'].str.lower().str.contains(item_name_lower, na=False, regex=False)
    matching = df[mask]
    
    if not matching.empty and year_col in matching.columns:
        return float(matching.iloc[0][year_col])
    return 0.0


def detect_screener_year_columns(df):
    """
    Detect year columns dynamically (columns starting with _)
    
    Returns:
        list: Sorted list of year column names
    """
    if df is None or df.empty:
        return []
    
    year_cols = [col for col in df.columns if col.startswith('_') and col != 'Item']
    # Sort by numeric value after underscore
    year_cols.sort(key=lambda x: int(x[1:]) if x[1:].isdigit() else 0)
    return year_cols


# ================================
# YAHOO FINANCE INTEGRATION
# ================================

def fetch_ticker_data_for_screener(ticker_symbol, exchange='NS'):
    """
    Fetch current price and beta from Yahoo Finance for Screener mode
    
    Args:
        ticker_symbol: Stock ticker (e.g., 'RELIANCE', 'TCS')
        exchange: Exchange suffix ('NS' for NSE, 'BO' for BSE)
    
    Returns:
        dict: {'current_price': float, 'beta': float, 'ticker': str}
    """
    if not ticker_symbol or not ticker_symbol.strip():
        return {'current_price': 0.0, 'beta': 1.0, 'ticker': None, 'error': 'No ticker provided'}
    
    try:
        # Add exchange suffix if not present
        ticker_clean = ticker_symbol.strip().upper()
        if '.NS' not in ticker_clean and '.BO' not in ticker_clean:
            full_ticker = f"{ticker_clean}.{exchange}"
        else:
            full_ticker = ticker_clean
        
        st.info(f"🔍 Fetching data for {full_ticker} from Yahoo Finance...")
        
        # Fetch ticker data
        ticker = safe_ticker(full_ticker)
        info = ticker.info
        
        # Get current price
        current_price = info.get('currentPrice', 0.0)
        if current_price == 0.0:
            current_price = info.get('regularMarketPrice', 0.0)
        if current_price == 0.0:
            current_price = info.get('previousClose', 0.0)
        
        # Get beta (still needed for WACC calculation, but not displayed)
        beta = info.get('beta', 1.0)
        if beta is None or beta <= 0:
            beta = 1.0

        # Shares outstanding -- NOT part of the original function (which only
        # fetched price+beta), added since Screener mode needs a Yahoo option
        # for shares alongside "from Excel" and "manual", same fallback chain
        # Listed mode already uses (sharesOutstanding, else marketCap/price).
        shares = info.get('sharesOutstanding', 0) or 0
        if not shares:
            market_cap = info.get('marketCap', 0) or 0
            if market_cap > 0 and current_price > 0:
                shares = int(market_cap / current_price)

        st.success(f"✅ Current Price: ₹{current_price:.2f}")
        
        return {
            'current_price': current_price,
            'beta': beta,
            'shares': shares,
            'ticker': full_ticker,
            'error': None
        }
        
    except Exception as e:
        st.error(f"❌ Error fetching ticker data: {str(e)}")
        return {
            'current_price': 0.0,
            'beta': 1.0,
            'ticker': ticker_symbol,
            'error': str(e)
        }


# ================================
# SCREENER FINANCIAL EXTRACTION
# ================================

def extract_screener_financials(df_bs, df_pl, year_cols):
    """
    Extract financial metrics from Screener Excel DataFrames
    
    **CRITICAL: Screener data is in CRORES, we convert to LACS (multiply by 100)**
    
    Maps Screener template items to financial metrics:
    
    Balance Sheet Items:
    - Equity Share Capital
    - Reserves
    - Borrowings (Debt)
    - Net Block (Fixed Assets)
    - Investments
    - Receivables
    - Inventory
    - Cash & Bank
    - No. of Equity Shares
    
    P&L Items:
    - Sales (Revenue)
    - Raw Material Cost, Change in Inventory, Power and Fuel, Other Mfr. Exp (COGS components)
    - Employee Cost, Selling and admin (OpEx)
    - Other Income, Other Expenses
    - Depreciation
    - Interest
    - Tax (ACTUAL TAX PAID, not PBT)
    - Net profit
    - Dividend Amount
    
    Returns:
        dict: Financial metrics with historical data (NEWEST FIRST) in LACS
        Also includes 'num_shares' extracted from Balance Sheet
    """
    # Use ALL provided year_cols (already filtered by user selection in main file)
    last_years = year_cols
    
    # REVERSE to match unlisted mode - NEWEST FIRST [0], OLDEST LAST [-1]
    last_years = list(reversed(last_years))
    
    financials = {
        'years': last_years,
        'revenue': [],
        'cogs': [],
        'opex': [],
        'ebitda': [],
        'depreciation': [],
        'ebit': [],
        'interest': [],
        'interest_income': [],  # For business classification
        'tax': [],
        'nopat': [],
        'net_profit': [],  # Actual reported net profit
        'dividends': [],  # Dividend amounts in CRORES (will convert to Lacs)
        'fixed_assets': [],
        'inventory': [],
        'receivables': [],
        'payables': [],
        'cash': [],
        'equity': [],
        'st_debt': [],
        'lt_debt': [],
        'num_shares': None  # Will be extracted from latest year
    }
    
    # Extract number of shares from latest year (FIRST in reversed list)
    latest_year = last_years[0]
    num_shares_raw = get_value_from_screener_df(df_bs, 'No. of Equity Shares', latest_year)
    if num_shares_raw > 0:
        financials['num_shares'] = int(num_shares_raw)
    
    for year_col in last_years:
        # ===== INCOME STATEMENT =====
        
        # Revenue (Sales) - IN CRORES
        revenue = get_value_from_screener_df(df_pl, 'Sales', year_col)
        
        # COGS Components - IN CRORES
        raw_material = get_value_from_screener_df(df_pl, 'Raw Material Cost', year_col)
        change_inventory = get_value_from_screener_df(df_pl, 'Change in Inventory', year_col)
        power_fuel = get_value_from_screener_df(df_pl, 'Power and Fuel', year_col)
        other_mfr = get_value_from_screener_df(df_pl, 'Other Mfr', year_col)
        
        # Note: Change in inventory can be negative (increase) or positive (decrease)
        # COGS = Raw Material + Change in Inventory + Power & Fuel + Other Manufacturing
        cogs = raw_material + change_inventory + power_fuel + other_mfr
        
        # Operating Expenses - IN CRORES
        employee_cost = get_value_from_screener_df(df_pl, 'Employee Cost', year_col)
        selling_admin = get_value_from_screener_df(df_pl, 'Selling and admin', year_col)
        
        opex = employee_cost + selling_admin
        
        # Other Income and Expenses - IN CRORES
        other_income = get_value_from_screener_df(df_pl, 'Other Income', year_col)
        other_expenses = get_value_from_screener_df(df_pl, 'Other Expenses', year_col)
        
        # Depreciation - IN CRORES
        depreciation = get_value_from_screener_df(df_pl, 'Depreciation', year_col)
        
        # Interest - IN CRORES
        interest = get_value_from_screener_df(df_pl, 'Interest', year_col)
        
        # Note: Screener template doesn't have separate interest income
        # If other_income is primarily interest income, it could be used for classification
        interest_income = 0.0  # Default to 0 unless template provides it
        
        # **CRITICAL FIX**: Tax - Extract ACTUAL TAX PAID (not PBT) - IN CRORES
        tax = get_value_from_screener_df(df_pl, 'Tax', year_col)
        
        # Net Profit (reported) - IN CRORES
        net_profit = get_value_from_screener_df(df_pl, 'Net profit', year_col)
        
        # Profit Before Tax - IN CRORES
        pbt = get_value_from_screener_df(df_pl, 'Profit before tax', year_col)
        
        # **CRITICAL FIX**: If Tax is not found or is zero, CALCULATE it from PBT and Net Profit
        if tax == 0 or abs(tax) < 0.01:
            if pbt > 0 and net_profit > 0:
                tax = pbt - net_profit  # Tax = PBT - PAT
                if tax < 0:
                    tax = 0  # Handle cases where net profit > PBT due to extraordinary items
        
        # Dividends - IN CRORES
        dividend_amount = get_value_from_screener_df(df_pl, 'Dividend Amount', year_col)
        
        # Calculate EBITDA and EBIT
        # EBITDA = Revenue - COGS - OpEx + Other Income - Other Expenses
        ebitda = revenue - cogs - opex + other_income - other_expenses
        
        # EBIT = EBITDA - Depreciation
        ebit = ebitda - depreciation
        
        # **CRITICAL FIX**: NOPAT calculation 
        # Instead of NOPAT = EBIT * (1 - tax_rate), use:
        # NOPAT = Net Profit + Interest * (1 - tax_rate)
        # This is more reliable when Excel has simplified P&L structure
        
        # Calculate effective tax rate from actual tax paid
        if pbt > 0 and abs(tax) > 0:
            effective_tax_rate = abs(tax) / pbt
        else:
            effective_tax_rate = 0.25  # Default 25%
        
        # NOPAT = PAT + Interest * (1 - Tax Rate)
        # This represents operating profit after tax, adding back interest (net of tax shield)
        nopat = net_profit + (interest * (1 - effective_tax_rate))
        
        # **CONVERT FROM CRORES TO LACS (multiply by 100)**
        financials['revenue'].append(revenue * 100)
        financials['cogs'].append(cogs * 100)
        financials['opex'].append(opex * 100)
        financials['ebitda'].append(ebitda * 100)
        financials['depreciation'].append(depreciation * 100)
        financials['ebit'].append(ebit * 100)
        financials['interest'].append(interest * 100)
        financials['interest_income'].append(interest_income * 100)
        financials['tax'].append(tax * 100)  # Store actual tax paid in Lacs
        financials['nopat'].append(nopat * 100)
        financials['net_profit'].append(net_profit * 100)
        financials['dividends'].append(dividend_amount * 100)  # Dividends also in Lacs
        
        # ===== BALANCE SHEET =====
        
        # Fixed Assets (Net Block + Capital Work in Progress) - IN CRORES
        net_block = get_value_from_screener_df(df_bs, 'Net Block', year_col)
        cwip = get_value_from_screener_df(df_bs, 'Capital Work in Progress', year_col)
        fixed_assets = net_block + cwip
        
        # Current Assets - IN CRORES
        inventory = get_value_from_screener_df(df_bs, 'Inventory', year_col)
        receivables = get_value_from_screener_df(df_bs, 'Receivables', year_col)
        cash = get_value_from_screener_df(df_bs, 'Cash & Bank', year_col)
        
        # Equity (Equity Share Capital + Reserves) - IN CRORES
        share_capital = get_value_from_screener_df(df_bs, 'Equity Share Capital', year_col)
        reserves = get_value_from_screener_df(df_bs, 'Reserves', year_col)
        equity = share_capital + reserves
        
        # Debt (Borrowings) - IN CRORES
        borrowings = get_value_from_screener_df(df_bs, 'Borrowings', year_col)
        # Screener template has only total borrowings, split as 60% LT, 40% ST
        lt_debt = borrowings * 0.6
        st_debt = borrowings * 0.4
        
        # Payables (from Other Liabilities - rough estimate) - IN CRORES
        other_liabilities = get_value_from_screener_df(df_bs, 'Other Liabilities', year_col)
        # Assume 70% of other liabilities are trade payables
        payables = other_liabilities * 0.7
        
        # **CONVERT FROM CRORES TO LACS (multiply by 100)**
        financials['fixed_assets'].append(fixed_assets * 100)
        financials['inventory'].append(inventory * 100)
        financials['receivables'].append(receivables * 100)
        financials['payables'].append(payables * 100)
        financials['cash'].append(cash * 100)
        financials['equity'].append(equity * 100)
        financials['st_debt'].append(st_debt * 100)
        financials['lt_debt'].append(lt_debt * 100)
    
    return financials


# ================================
# SCREENER SPECIFIC UTILITIES
# ================================

def get_screener_shares_outstanding(df_bs, year_col):
    """
    Extract number of shares outstanding from Balance Sheet
    
    Returns:
        int: Number of shares
    """
    shares = get_value_from_screener_df(df_bs, 'No. of Equity Shares', year_col)
    if shares > 0:
        return int(shares)
    
    # Try without the period
    shares = get_value_from_screener_df(df_bs, 'No of Equity Shares', year_col)
    if shares > 0:
        return int(shares)
    
    return 0


def get_screener_face_value(df_bs, year_col):
    """
    Extract face value per share from Balance Sheet
    
    Returns:
        float: Face value
    """
    face_value = get_value_from_screener_df(df_bs, 'Face value', year_col)
    return face_value if face_value > 0 else 10.0  # Default to 10


# ================================
# DDM (DIVIDEND DISCOUNT MODEL) FOR SCREENER
# ================================

def calculate_screener_ddm_valuation(financials, num_shares, required_return=0.12, growth_rate=0.05):
    """
    Calculate DDM valuation using historical dividend data from Screener template
    
    Args:
        financials: Dict with 'dividends' key containing historical dividend amounts
        num_shares: Number of shares outstanding
        required_return: Required rate of return (default 12%)
        growth_rate: Expected dividend growth rate (default 5%)
    
    Returns:
        dict: DDM valuation results
    """
    if not financials or 'dividends' not in financials or num_shares <= 0:
        return None
    
    # Convert None/NaN to 0 for blank cells
    dividends = [float(d) if d and not (isinstance(d, float) and np.isnan(d)) else 0.0 for d in financials['dividends']]
    years = financials.get('years', [])
    
    # Filter out zero dividends for analysis
    non_zero_divs = [d for d in dividends if d > 0]
    
    if len(non_zero_divs) < 2:
        return {
            'model': 'DDM',
            'status': 'Insufficient Data',
            'message': 'Need at least 2 years of dividend history',
            'value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    # Calculate historical dividend growth rate (only for non-zero consecutive years)
    historical_growth_rates = []
    for i in range(1, len(dividends)):
        if dividends[i-1] > 0 and dividends[i] > 0:
            growth = (dividends[i] - dividends[i-1]) / dividends[i-1]
            historical_growth_rates.append(growth)
    
    avg_historical_growth = np.mean(historical_growth_rates) if historical_growth_rates else growth_rate
    
    # Use provided growth_rate if it's not the default (0.05), otherwise use historical
    # If growth_rate is 0.05 (default), use historical growth
    # If growth_rate is different, user has overridden it - use that
    if abs(growth_rate - 0.05) > 0.001:  # User has provided a custom growth rate
        final_growth_rate = growth_rate
    else:  # Use historical growth
        final_growth_rate = avg_historical_growth
    
    # Use latest non-zero dividend (in LACS)
    latest_dividend = next((d for d in reversed(dividends) if d > 0), 0)
    
    if latest_dividend == 0:
        return {
            'model': 'DDM',
            'status': 'No Recent Dividend',
            'message': 'No dividend paid in recent years',
            'value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    # Convert dividend from LACS to RUPEES, then divide by shares
    dps = (latest_dividend * 100000) / num_shares  # Dividend per share in Rupees
    
    # Gordon Growth Model: P = D1 / (r - g)
    # D1 = D0 * (1 + g)
    if required_return <= avg_historical_growth:
        return {
            'model': 'DDM (Gordon Growth Model)',
            'status': 'Invalid',
            'message': f'Required return ({required_return*100:.1f}%) must be greater than growth rate ({avg_historical_growth*100:.1f}%)',
            'value_per_share': 0,
            'total_intrinsic_value': 0,
            'latest_dps': dps,
            'historical_growth_rate': avg_historical_growth
        }
    
    # Calculate next year's expected dividend
    d1 = dps * (1 + avg_historical_growth)
    
    # Intrinsic value per share
    intrinsic_value_per_share = d1 / (required_return - avg_historical_growth)
    
    # Total intrinsic value
    total_intrinsic_value = intrinsic_value_per_share * num_shares
    
    return {
        'model': 'DDM (Gordon Growth Model)',
        'status': 'Success',
        'latest_dividend_total': latest_dividend,
        'latest_dps': dps,
        'num_shares': num_shares,
        'historical_growth_rate': avg_historical_growth,
        'assumed_growth_rate': avg_historical_growth,
        'required_return': required_return,
        'expected_next_dividend': d1,
        'value_per_share': intrinsic_value_per_share,
        'total_intrinsic_value': total_intrinsic_value,
        'dividend_history': dividends,
        'years': years
    }


# ================================
# RIM (RESIDUAL INCOME MODEL) FOR SCREENER
# ================================

def calculate_screener_rim_valuation(financials, num_shares, required_return=0.12, projection_years=5, terminal_growth=0.04, assumed_roe=None):
    """
    Calculate RIM (Residual Income Model) valuation
    
    RIM Formula:
    Value = Book Value + PV(Expected Residual Income)
    Residual Income = Net Income - (Book Value * Required Return)
    
    Args:
        financials: Dict with historical financial data
        num_shares: Number of shares outstanding
        required_return: Required rate of return (default 12%)
        projection_years: Number of years to project (default 5)
        terminal_growth: Terminal growth rate (default 4%)
        assumed_roe: Override ROE assumption (None = auto-calculate from historical data)
    
    Returns:
        dict: RIM valuation results
    """
    if not financials or num_shares <= 0:
        return None
    
    # Get latest book value and net income (both in LACS)
    equity = financials['equity'][0]  # Total equity (book value) in LACS - index 0 is NEWEST
    net_income = financials['net_profit'][0]  # Use reported net profit in LACS
    
    if equity <= 0:
        return {
            'model': 'RIM',
            'status': 'Invalid',
            'message': 'Book value must be positive',
            'value_per_share': 0,
            'total_intrinsic_value': 0
        }
    
    # Calculate historical ROE and growth rate
    if len(financials['equity']) >= 2 and len(financials['net_profit']) >= 2:
        roe_values = []
        for i in range(len(financials['equity'])):
            if financials['equity'][i] > 0:
                roe = financials['net_profit'][i] / financials['equity'][i]
                roe_values.append(roe)
        avg_historical_roe = np.mean(roe_values) if roe_values else 0.15  # Default 15%
        
        # Calculate earnings growth rate
        earnings_growth_rates = []
        for i in range(1, len(financials['net_profit'])):
            if financials['net_profit'][i-1] > 0:
                growth = (financials['net_profit'][i] - financials['net_profit'][i-1]) / financials['net_profit'][i-1]
                earnings_growth_rates.append(growth)
        avg_earnings_growth = np.mean(earnings_growth_rates) if earnings_growth_rates else 0.08  # Default 8%
    else:
        avg_historical_roe = 0.15
        avg_earnings_growth = 0.08
    
    # Use assumed_roe if provided, otherwise use historical
    avg_roe = assumed_roe if assumed_roe is not None and assumed_roe > 0 else avg_historical_roe
    
    # Project residual income
    residual_incomes = []
    projected_book_values = [equity]
    projected_earnings = [net_income]
    
    for year in range(1, projection_years + 1):
        # Project next year's book value and earnings
        next_book_value = projected_book_values[-1] * (1 + avg_earnings_growth)
        next_earnings = projected_earnings[-1] * (1 + avg_earnings_growth)
        
        # Calculate residual income
        # RI = Net Income - (Required Return × Book Value)
        required_income = projected_book_values[-1] * required_return
        residual_income = next_earnings - required_income
        
        residual_incomes.append(residual_income)
        projected_book_values.append(next_book_value)
        projected_earnings.append(next_earnings)
    
    # Calculate present value of residual incomes
    pv_residual_incomes = []
    for i, ri in enumerate(residual_incomes, 1):
        pv = ri / ((1 + required_return) ** i)
        pv_residual_incomes.append(pv)
    
    # Terminal value (perpetuity)
    if required_return > terminal_growth:
        terminal_ri = residual_incomes[-1] * (1 + terminal_growth)
        terminal_value = terminal_ri / (required_return - terminal_growth)
        pv_terminal = terminal_value / ((1 + required_return) ** projection_years)
    else:
        pv_terminal = 0
    
    # Total intrinsic value (in LACS)
    # Value = Current Book Value + PV(Residual Incomes) + PV(Terminal Value)
    total_intrinsic_value = equity + sum(pv_residual_incomes) + pv_terminal
    
    # Convert from LACS to RUPEES for per share calculations
    total_intrinsic_value_rupees = total_intrinsic_value * 100000
    intrinsic_value_per_share = total_intrinsic_value_rupees / num_shares
    
    return {
        'model': 'RIM (Residual Income Model)',
        'status': 'Success',
        'book_value': equity,
        'net_income': net_income,
        'avg_roe': avg_roe,
        'avg_earnings_growth': avg_earnings_growth,
        'required_return': required_return,
        'projection_years': projection_years,
        'terminal_growth': terminal_growth,
        'residual_incomes': residual_incomes,
        'pv_residual_incomes': pv_residual_incomes,
        'pv_terminal': pv_terminal,
        'total_intrinsic_value': total_intrinsic_value,
        'value_per_share': intrinsic_value_per_share,
        'num_shares': num_shares
    }


# ================================
# EXCEL GENERATION FOR SCREENER MODE
# ================================

def generate_screener_valuation_excel(company_name, financials, dcf_results, ddm_results, rim_results, comp_val_results, peer_comparison):
    """
    Generate comprehensive Excel report for Screener mode valuation
    
    Args:
        company_name: Company name
        financials: Historical financial data dict
        dcf_results: DCF valuation results dict
        ddm_results: DDM valuation results dict
        rim_results: RIM valuation results dict
        comp_val_results: Comparative valuation results dict
        peer_comparison: Peer comparison data (not used in Screener mode)
    
    Returns:
        BytesIO: Excel file buffer
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    subheader_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: Summary =====
    ws_summary = wb.create_sheet('Summary')
    
    ws_summary['A1'] = f'{company_name} - Valuation Summary'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    row = 3
    ws_summary[f'A{row}'] = 'Valuation Method'
    ws_summary[f'B{row}'] = 'Fair Value per Share (₹)'
    ws_summary[f'C{row}'] = 'Status'
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'C{row}'].fill = header_fill
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'B{row}'].font = header_font
    ws_summary[f'C{row}'].font = header_font
    
    row += 1
    
    # DCF
    if dcf_results:
        ws_summary[f'A{row}'] = 'DCF (FCFF)'
        ws_summary[f'B{row}'] = dcf_results.get('fair_value_per_share', 0)
        ws_summary[f'B{row}'].number_format = '₹#,##0.00'
        ws_summary[f'C{row}'] = 'Complete'
        row += 1
    
    # DDM
    if ddm_results and ddm_results.get('status') == 'Success':
        ws_summary[f'A{row}'] = 'DDM (Gordon Growth)'
        ws_summary[f'B{row}'] = ddm_results.get('value_per_share', 0)
        ws_summary[f'B{row}'].number_format = '₹#,##0.00'
        ws_summary[f'C{row}'] = ddm_results.get('status', '')
        row += 1
    
    # RIM
    if rim_results and rim_results.get('status') == 'Success':
        ws_summary[f'A{row}'] = 'RIM (Residual Income)'
        ws_summary[f'B{row}'] = rim_results.get('value_per_share', 0)
        ws_summary[f'B{row}'].number_format = '₹#,##0.00'
        ws_summary[f'C{row}'] = rim_results.get('status', '')
        row += 1
    
    # Comparative Valuation
    if comp_val_results:
        # Calculate average of all comparative methods
        valuations = comp_val_results.get('valuations', {})
        if valuations:
            all_avg_values = [v.get('fair_value_avg', 0) for v in valuations.values()]
            avg_comp_value = np.mean(all_avg_values) if all_avg_values else 0
            
            ws_summary[f'A{row}'] = 'Comparative Valuation (Avg)'
            ws_summary[f'B{row}'] = avg_comp_value
            ws_summary[f'B{row}'].number_format = '₹#,##0.00'
            ws_summary[f'C{row}'] = 'Complete'
            row += 1
    
    # Auto-fit columns
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 15
    
    # ===== SHEET 2: Historical Financials =====
    ws_hist = wb.create_sheet('Historical Financials')
    
    ws_hist['A1'] = 'Historical Financial Data (All values in ₹ Lacs)'
    ws_hist['A1'].font = Font(bold=True, size=12)
    ws_hist.merge_cells('A1:E1')
    
    row = 3
    ws_hist[f'A{row}'] = 'Metric'
    for i, year in enumerate(financials['years']):
        ws_hist[f'{chr(66+i)}{row}'] = str(year)
        ws_hist[f'{chr(66+i)}{row}'].fill = header_fill
        ws_hist[f'{chr(66+i)}{row}'].font = header_font
    ws_hist[f'A{row}'].fill = header_fill
    ws_hist[f'A{row}'].font = header_font
    
    # Income Statement
    row += 1
    metrics = [
        ('Revenue', 'revenue'),
        ('COGS', 'cogs'),
        ('Operating Expenses', 'opex'),
        ('EBITDA', 'ebitda'),
        ('Depreciation', 'depreciation'),
        ('EBIT', 'ebit'),
        ('Interest', 'interest'),
        ('Tax', 'tax'),
        ('NOPAT', 'nopat'),
        ('Net Profit', 'net_profit')
    ]
    
    for label, key in metrics:
        ws_hist[f'A{row}'] = label
        for i, val in enumerate(financials.get(key, [])):
            ws_hist[f'{chr(66+i)}{row}'] = val
            ws_hist[f'{chr(66+i)}{row}'].number_format = '#,##0.00'
        row += 1
    
    # Balance Sheet
    row += 1
    ws_hist[f'A{row}'] = 'Balance Sheet'
    ws_hist[f'A{row}'].fill = subheader_fill
    ws_hist[f'A{row}'].font = subheader_font
    ws_hist.merge_cells(f'A{row}:E{row}')
    
    row += 1
    bs_metrics = [
        ('Fixed Assets', 'fixed_assets'),
        ('Inventory', 'inventory'),
        ('Receivables', 'receivables'),
        ('Payables', 'payables'),
        ('Cash', 'cash'),
        ('Equity', 'equity'),
        ('ST Debt', 'st_debt'),
        ('LT Debt', 'lt_debt')
    ]
    
    for label, key in bs_metrics:
        ws_hist[f'A{row}'] = label
        for i, val in enumerate(financials.get(key, [])):
            ws_hist[f'{chr(66+i)}{row}'] = val
            ws_hist[f'{chr(66+i)}{row}'].number_format = '#,##0.00'
        row += 1
    
    ws_hist.column_dimensions['A'].width = 25
    for i in range(len(financials['years'])):
        ws_hist.column_dimensions[chr(66+i)].width = 15
    
    # ===== SHEET 3: DCF Results =====
    if dcf_results:
        ws_dcf = wb.create_sheet('DCF Valuation')
        
        ws_dcf['A1'] = 'DCF (FCFF) Valuation Results'
        ws_dcf['A1'].font = Font(bold=True, size=12)
        ws_dcf.merge_cells('A1:D1')
        
        row = 3
        ws_dcf[f'A{row}'] = 'Metric'
        ws_dcf[f'B{row}'] = 'Value'
        ws_dcf[f'A{row}'].fill = header_fill
        ws_dcf[f'B{row}'].fill = header_fill
        ws_dcf[f'A{row}'].font = header_font
        ws_dcf[f'B{row}'].font = header_font
        
        row += 1
        dcf_metrics = [
            ('Enterprise Value', dcf_results.get('enterprise_value', 0), '₹#,##0.00 Lacs'),
            ('Total Debt', dcf_results.get('total_debt', 0), '₹#,##0.00 Lacs'),
            ('Cash & Equivalents', dcf_results.get('cash', 0), '₹#,##0.00 Lacs'),
            ('Equity Value', dcf_results.get('equity_value', 0), '₹#,##0.00 Lacs'),
            ('Number of Shares', dcf_results.get('shares', financials.get('num_shares', 0)), '#,##0'),
            ('Fair Value per Share', dcf_results.get('fair_value_per_share', 0), '₹#,##0.00'),
            ('WACC', dcf_results.get('wacc', 0), '0.00%'),
            ('Terminal Value %', dcf_results.get('tv_percentage', 0), '0.0%')
        ]
        
        for label, value, fmt in dcf_metrics:
            ws_dcf[f'A{row}'] = label
            ws_dcf[f'B{row}'] = value
            ws_dcf[f'B{row}'].number_format = fmt
            row += 1
        
        ws_dcf.column_dimensions['A'].width = 30
        ws_dcf.column_dimensions['B'].width = 20
    
    # ===== SHEET 4: DDM Results =====
    if ddm_results and ddm_results.get('status') == 'Success':
        ws_ddm = wb.create_sheet('DDM Valuation')
        
        ws_ddm['A1'] = 'Dividend Discount Model (Gordon Growth)'
        ws_ddm['A1'].font = Font(bold=True, size=12)
        ws_ddm.merge_cells('A1:D1')
        
        row = 3
        ws_ddm[f'A{row}'] = 'Parameter'
        ws_ddm[f'B{row}'] = 'Value'
        ws_ddm[f'A{row}'].fill = header_fill
        ws_ddm[f'B{row}'].fill = header_fill
        ws_ddm[f'A{row}'].font = header_font
        ws_ddm[f'B{row}'].font = header_font
        
        row += 1
        ddm_metrics = [
            ('Latest DPS', ddm_results.get('latest_dps', 0), '₹#,##0.00'),
            ('Historical Growth Rate', ddm_results.get('historical_growth_rate', 0)*100, '0.00%'),
            ('Required Return', ddm_results.get('required_return', 0)*100, '0.00%'),
            ('Expected Next Dividend', ddm_results.get('expected_next_dividend', 0), '₹#,##0.00'),
            ('Fair Value per Share', ddm_results.get('value_per_share', 0), '₹#,##0.00')
        ]
        
        for label, value, fmt in ddm_metrics:
            ws_ddm[f'A{row}'] = label
            ws_ddm[f'B{row}'] = value
            ws_ddm[f'B{row}'].number_format = fmt
            row += 1
        
        ws_ddm.column_dimensions['A'].width = 30
        ws_ddm.column_dimensions['B'].width = 20
    
    # ===== SHEET 5: RIM Results =====
    if rim_results and rim_results.get('status') == 'Success':
        ws_rim = wb.create_sheet('RIM Valuation')
        
        ws_rim['A1'] = 'Residual Income Model'
        ws_rim['A1'].font = Font(bold=True, size=12)
        ws_rim.merge_cells('A1:D1')
        
        row = 3
        ws_rim[f'A{row}'] = 'Parameter'
        ws_rim[f'B{row}'] = 'Value'
        ws_rim[f'A{row}'].fill = header_fill
        ws_rim[f'B{row}'].fill = header_fill
        ws_rim[f'A{row}'].font = header_font
        ws_rim[f'B{row}'].font = header_font
        
        row += 1
        rim_metrics = [
            ('Book Value (Equity)', rim_results.get('book_value', 0), '₹#,##0.00 Lacs'),
            ('Net Income', rim_results.get('net_income', 0), '₹#,##0.00 Lacs'),
            ('Average ROE', rim_results.get('avg_roe', 0)*100, '0.00%'),
            ('Average Earnings Growth', rim_results.get('avg_earnings_growth', 0)*100, '0.00%'),
            ('Required Return', rim_results.get('required_return', 0)*100, '0.00%'),
            ('Terminal Growth', rim_results.get('terminal_growth', 0)*100, '0.00%'),
            ('Fair Value per Share', rim_results.get('value_per_share', 0), '₹#,##0.00')
        ]
        
        for label, value, fmt in rim_metrics:
            ws_rim[f'A{row}'] = label
            ws_rim[f'B{row}'] = value
            ws_rim[f'B{row}'].number_format = fmt
            row += 1
        
        ws_rim.column_dimensions['A'].width = 30
        ws_rim.column_dimensions['B'].width = 20
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def calculate_manual_peer_valuation(financials, shares, peers):
    """Simple average P/E and EV/EBITDA multiple valuation from manually
    entered peer data (name + P/E and/or EV/EBITDA multiple) — used by
    Screener Excel Mode's manual peer comparison section, since Screener
    exports don't come with a ticker to auto-fetch real peer financials."""
    import numpy as _np

    pe_multiples = [p["pe"] for p in peers if p.get("pe", 0) > 0]
    ev_ebitda_multiples = [p["ev_ebitda"] for p in peers if p.get("ev_ebitda", 0) > 0]

    avg_pe = _np.mean(pe_multiples) if pe_multiples else 0
    avg_ev_ebitda = _np.mean(ev_ebitda_multiples) if ev_ebitda_multiples else 0

    latest_net_profit = financials.get("nopat", [0])[0]
    eps = (latest_net_profit * 100000) / shares if shares > 0 else 0
    pe_value_per_share = eps * avg_pe

    latest_ebitda = financials.get("ebitda", [0])[0]
    enterprise_value = latest_ebitda * avg_ev_ebitda
    net_debt = (financials.get("st_debt", [0])[0] + financials.get("lt_debt", [0])[0]) - financials.get("cash", [0])[0]
    equity_value = enterprise_value - net_debt
    ev_ebitda_value_per_share = (equity_value * 100000) / shares if shares > 0 else 0

    values = [v for v in (pe_value_per_share, ev_ebitda_value_per_share) if v > 0]
    avg_value = _np.mean(values) if values else 0

    return {
        "pe_multiple": avg_pe,
        "ev_ebitda_multiple": avg_ev_ebitda,
        "pe_value": pe_value_per_share,
        "ev_ebitda_value": ev_ebitda_value_per_share,
        "avg_value": avg_value,
        "peers": peers,
    }


# ================================
# DISPLAY FUNCTIONS FOR STREAMLIT
# ================================

