"""
Screener.in Auto Downloader with Excel Converter
=================================================
Downloads Excel from Screener.in, removes blank columns, converts to template format
"""

import requests
from bs4 import BeautifulSoup
import pickle
import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.logic.yf_ratelimit import _make_session


class ScreenerDownloader:
    """Downloads Excel files from Screener.in with authentication"""
    
    def __init__(self, cookies_path="screener_cookies.pkl"):
        """
        Initialize downloader with cookies
        
        Args:
            cookies_path: Path to pickled cookies file
        """
        self.cookies_path = cookies_path
        # Chrome-impersonation session (same fix that resolved Yahoo Finance's
        # "'str' object has no attribute 'name'" blocking) — plain
        # requests.Session() was getting a 403 from Screener.in in testing.
        self.session = _make_session()
        self._load_cookies()
        
    def _load_cookies(self):
        """Load cookies from a pickle (.pkl) or JSON (.json) file.

        JSON support added for deployment platforms like Render where
        "Secret Files" are pasted as plaintext — a binary pickle file
        doesn't survive that round trip, but a JSON list of
        {"name": ..., "value": ...} cookie dicts does. Same content,
        just a text-safe format.
        """
        if not os.path.exists(self.cookies_path):
            raise FileNotFoundError(f"Cookies file not found: {self.cookies_path}")

        if str(self.cookies_path).lower().endswith(".json"):
            import json
            with open(self.cookies_path, "r") as f:
                cookies = json.load(f)
        else:
            with open(self.cookies_path, 'rb') as f:
                cookies = pickle.load(f)
        
        # Convert to requests cookies
        if isinstance(cookies, list):
            for cookie in cookies:
                self.session.cookies.set(cookie['name'], cookie['value'])
        elif isinstance(cookies, dict):
            for name, value in cookies.items():
                self.session.cookies.set(name, value)
    
    def download_excel(self, company_symbol, output_path=None, use_consolidated=False, use_id_url=False):
        """
        Download Excel file from Screener.in by clicking Export button
        
        Args:
            company_symbol: Company symbol (e.g., 'HONASA') or ID number (e.g., '1285886')
            output_path: Where to save the file (optional)
            use_consolidated: Use consolidated financials (default: False)
            use_id_url: Use ID-based URL format /company/id/NUMBER/ (default: False)
            
        Returns:
            str: Path to downloaded file or None if failed
        """
        # Construct URL based on flags
        if use_id_url:
            url_suffix = "consolidated/" if use_consolidated else ""
            company_url = f"https://www.screener.in/company/id/{company_symbol}/{url_suffix}"
        else:
            url_suffix = "consolidated/" if use_consolidated else ""
            company_url = f"https://www.screener.in/company/{company_symbol}/{url_suffix}"
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Referer': 'https://www.screener.in/',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        # Set up session with retries for Streamlit Cloud compatibility
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
        import os
        
        # Disable proxy if it's blocking screener.in
        self.session.trust_env = False
        os.environ.pop('HTTP_PROXY', None)
        os.environ.pop('HTTPS_PROXY', None)
        os.environ.pop('http_proxy', None)
        os.environ.pop('https_proxy', None)
        
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        try:
            self.session.mount("https://", adapter)
            self.session.mount("http://", adapter)
        except AttributeError:
            pass  # curl_cffi sessions don't support .mount() — retries aren't critical here
        
        try:
            print(f"Accessing: {company_url}")
            print(f"Environment: Checking Streamlit Cloud compatibility...")
            
            # Try with verify=True first (proper SSL)
            try:
                response = self.session.get(company_url, headers=headers, timeout=30, verify=True)
            except requests.exceptions.SSLError:
                print("SSL verification failed, trying without SSL verification...")
                response = self.session.get(company_url, headers=headers, timeout=30, verify=False)
                
        except requests.exceptions.ConnectionError as e:
            error_msg = str(e)
            print(f"❌ CONNECTION ERROR: Cannot reach www.screener.in")
            print(f"Error details: {error_msg}")
            
            # Check if it's specifically Streamlit Cloud issue
            if "Connection refused" in error_msg or "Errno 111" in error_msg or "Proxy" in error_msg or "403 Forbidden" in error_msg:
                print(f"\n🔴 STREAMLIT CLOUD NETWORK RESTRICTION")
                print(f"⚠️  **RECENT CHANGE**: Streamlit Cloud recently blocked access to www.screener.in")
                print(f"This is a platform-level restriction that was added after your app was working.")
                print(f"\n✅ RECOMMENDED SOLUTIONS:")
                print(f"1. **Use Screener Excel Mode**: Upload manually downloaded Excel files")
                print(f"   - Go to www.screener.in/company/{company_symbol}/consolidated/")
                print(f"   - Click 'Export' button to download Excel")
                print(f"   - Upload the file in the app's 'Screener Excel Mode'")
                print(f"\n2. **Deploy on Different Platform**: Use Heroku, Railway, or Render (free options)")
                print(f"\n3. **Use Yahoo Finance mode**: For listed companies with NSE/BSE tickers")
            else:
                print(f"\n⚠️  Network connection issue")
                print(f"Possible causes:")
                print(f"- Firewall blocking the connection")
                print(f"- DNS resolution failure")
                print(f"- Streamlit Cloud network policies")
            
            return None
            
        except requests.exceptions.Timeout:
            print(f"❌ TIMEOUT: Request to www.screener.in timed out after 30 seconds")
            print(f"The server may be slow or your network connection is unstable.")
            print(f"Try again later or use the Excel upload feature.")
            return None
            
        except Exception as e:
            print(f"❌ UNEXPECTED ERROR: {type(e).__name__}: {str(e)}")
            import traceback
            print(f"Full traceback:")
            traceback.print_exc()
            return None
        
        # Process the response
        try:
            if response.status_code != 200:
                print(f"Error: Could not access page (Status: {response.status_code})")
                if response.status_code == 403:
                    print(f"⚠️  403 Forbidden - Access denied by server. Authentication may be required.")
                elif response.status_code == 404:
                    print(f"⚠️  404 Not Found - Company '{company_symbol}' not found on Screener.in")
                elif response.status_code == 429:
                    print(f"⚠️  429 Too Many Requests - Rate limited. Wait and try again.")
                return None
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find export button
            export_button = soup.find('button', {'formaction': re.compile(r'/user/company/export/\d+/')})
            
            if not export_button:
                print("Error: Could not find export button")
                return None
            
            formaction = export_button.get('formaction')
            print(f"Found export URL: {formaction}")
            
            # Get CSRF token
            form = export_button.find_parent('form')
            csrf_token = None
            if form:
                csrf_input = form.find('input', {'name': 'csrfmiddlewaretoken'})
                if csrf_input:
                    csrf_token = csrf_input.get('value')
            
            if not csrf_token:
                csrf_token = self.session.cookies.get('csrftoken', '')
            
            # POST to export URL
            export_url = f"https://www.screener.in{formaction}"
            
            post_headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
                'Referer': company_url,
                'Origin': 'https://www.screener.in'
            }
            
            post_data = {
                'csrfmiddlewaretoken': csrf_token,
                'next': f'/company/id/{company_symbol}/{url_suffix}' if use_id_url else f'/company/{company_symbol}/{url_suffix}'
            }
            
            print(f"Downloading from: {export_url}")
            download_response = self.session.post(export_url, headers=post_headers, data=post_data, timeout=30)
            
            if download_response.status_code != 200:
                print(f"Error: Download failed (Status: {download_response.status_code})")
                return None
            
            # --- Validate content is a real Excel file before saving ---
            content = download_response.content
            content_type = download_response.headers.get('Content-Type', '')
            
            # XLSX files are ZIP archives — they must start with PK magic bytes (0x50 0x4B)
            is_xlsx = len(content) > 4 and content[:2] == b'PK'
            
            if not is_xlsx:
                snippet = content[:500].decode('utf-8', errors='replace')
                print(f"Error: Downloaded content is not a valid Excel file.")
                print(f"Content-Type received: {content_type}")
                print(f"First 500 chars of response:\n{snippet}")
                
                if 'login' in snippet.lower() or 'sign in' in snippet.lower() or 'password' in snippet.lower():
                    print("\n⚠️  Authentication failure — Screener.in returned a login page.")
                    print("Your cookies have likely expired. Please refresh screener_cookies.pkl.")
                elif 'csrf' in snippet.lower() or 'forbidden' in snippet.lower():
                    print("\n⚠️  CSRF/permission error. Try refreshing cookies and retrying.")
                elif '<html' in snippet.lower():
                    print("\n⚠️  Server returned an HTML page instead of the Excel file.")
                    print("Possible causes: rate limiting, session expiry, or the export URL changed.")
                else:
                    print("\n⚠️  Unexpected response format — not an Excel file.")
                return None
            
            # Save file
            if output_path is None:
                output_path = f"{company_symbol}_screener.xlsx"
            
            with open(output_path, 'wb') as f:
                f.write(content)
            
            if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                print(f"✓ Downloaded: {output_path} ({os.path.getsize(output_path)} bytes)")
                return output_path
            else:
                print("Error: File empty or not created")
                return None
                
        except Exception as e:
            print(f"Error downloading: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def remove_empty_year_columns(self, excel_path):
        """
        Remove columns from Data Sheet that don't have data in financial rows
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            bool: True if successful
        """
        try:
            # Validate file is a real Excel (ZIP) file before opening
            with open(excel_path, 'rb') as _f:
                magic = _f.read(2)
            if magic != b'PK':
                print(f"Error removing empty columns: File is not a valid Excel file (not a zip/xlsx). "
                      f"Got magic bytes: {magic!r}. The downloaded file may be an HTML page or error response.")
                return False
            
            wb = load_workbook(excel_path)
            
            if 'Data Sheet' not in wb.sheetnames:
                print("Data Sheet not found")
                wb.close()
                return False
            
            ws = wb['Data Sheet']
            
            # Find P&L section
            pl_date_row = None
            for i in range(1, 50):
                val = ws.cell(i, 1).value
                if val and 'Report Date' in str(val):
                    pl_date_row = i
                    break
            
            if not pl_date_row:
                print("Could not find Report Date row")
                wb.close()
                return False
            
            # Check which columns have actual year data
            cols_to_delete = []
            for col in range(2, ws.max_column + 1):
                date_val = ws.cell(pl_date_row, col).value
                
                # Check if this column has a valid date
                has_valid_date = False
                if date_val:
                    if hasattr(date_val, 'year'):
                        has_valid_date = True
                    else:
                        if re.search(r'20\d{2}', str(date_val)):
                            has_valid_date = True
                
                if not has_valid_date:
                    cols_to_delete.append(col)
                    continue
                
                # Also check if column has any financial data (check Sales row)
                sales_row = pl_date_row + 1
                has_data = False
                for check_row in range(sales_row, min(sales_row + 10, ws.max_row + 1)):
                    val = ws.cell(check_row, col).value
                    if val and val != 0:
                        try:
                            float(val)
                            has_data = True
                            break
                        except:
                            pass
                
                if not has_data:
                    cols_to_delete.append(col)
            
            # Delete columns in reverse order
            for col in sorted(cols_to_delete, reverse=True):
                ws.delete_cols(col)
                print(f"✓ Deleted empty column {col}")
            
            if cols_to_delete:
                wb.save(excel_path)
                print(f"✓ Removed {len(cols_to_delete)} empty columns")
            else:
                print("No empty columns to remove")
            
            wb.close()
            return True
            
        except Exception as e:
            print(f"Error removing empty columns: {e}")
            return False
    
    def remove_blank_columns(self, excel_path):
        """Legacy method - calls remove_empty_year_columns"""
        return self.remove_empty_year_columns(excel_path)
    
    def convert_to_template(self, screener_excel_path, output_path=None):
        """
        Convert Screener Data Sheet to EXACT target format
        
        Target format:
        - Sheet 1: "Balance Sheet" with title row, Report Date row, then data
        - Sheet 2: "Profit and Loss Account" with title row, Report Date row, then data
        """
        try:
            from openpyxl import load_workbook, Workbook
            
            # Validate file is a real Excel (ZIP) file before opening
            with open(screener_excel_path, 'rb') as _f:
                magic = _f.read(2)
            if magic != b'PK':
                print(f"Error converting: File is not a valid Excel file (not a zip/xlsx). "
                      f"Got magic bytes: {magic!r}. The file may be an HTML page or error response.")
                return None
            
            # Load source file
            src_wb = load_workbook(screener_excel_path, data_only=True)
            
            if 'Data Sheet' not in src_wb.sheetnames:
                print("Error: Data Sheet not found")
                src_wb.close()
                return None
            
            src_ws = src_wb['Data Sheet']
            
            # Find sections in Data Sheet
            pl_date_row = None
            bs_date_row = None
            
            for i in range(1, 100):
                val = src_ws.cell(i, 1).value
                if val:
                    val_str = str(val).upper()
                    if ('PROFIT' in val_str or 'P&L' in val_str or 'P & L' in val_str) and pl_date_row is None:
                        # Next row is Report Date
                        if src_ws.cell(i + 1, 1).value and 'Report Date' in str(src_ws.cell(i + 1, 1).value):
                            pl_date_row = i + 1
                    elif 'BALANCE' in val_str and bs_date_row is None:
                        if src_ws.cell(i + 1, 1).value and 'Report Date' in str(src_ws.cell(i + 1, 1).value):
                            bs_date_row = i + 1
            
            if not pl_date_row or not bs_date_row:
                print(f"Error: Sections not found. PL:{pl_date_row}, BS:{bs_date_row}")
                src_wb.close()
                return None
            
            print(f"P&L date row: {pl_date_row}, BS date row: {bs_date_row}")
            
            # Find first column with data (skip empty columns at start)
            first_data_col = None
            for col in range(2, src_ws.max_column + 1):
                val = src_ws.cell(pl_date_row, col).value
                if val:
                    first_data_col = col
                    break
            
            if not first_data_col:
                print("Error: No data columns found")
                src_wb.close()
                return None
            
            # Extract all dates/years from Report Date row
            dates = []
            date_cols = []
            for col in range(first_data_col, src_ws.max_column + 1):
                val = src_ws.cell(pl_date_row, col).value
                if val:
                    dates.append(val)
                    date_cols.append(col)
            
            if not dates:
                print("Error: No dates found")
                src_wb.close()
                return None
            
            print(f"✓ Found {len(dates)} date columns starting at column {first_data_col}")
            
            # Create new workbook
            new_wb = Workbook()
            new_wb.remove(new_wb.active)  # Remove default sheet
            
            # ============== BALANCE SHEET (First Sheet) ==============
            bs_ws = new_wb.create_sheet("Balance Sheet", 0)
            
            # Row 1: Title
            bs_ws['A1'] = 'BALANCE SHEET'
            
            # Row 2: Report Date header + dates
            bs_ws['A2'] = 'Report Date'
            for idx, date_val in enumerate(dates, start=2):
                bs_ws.cell(2, idx).value = date_val
            
            # Define Balance Sheet items IN EXACT ORDER from target
            bs_items = [
                'Equity Share Capital',
                'Reserves',
                'Borrowings',
                'Other Liabilities',
                'Total',  # First Total = Total Liabilities
                'Net Block',
                'Capital Work in Progress',
                'Investments',
                'Other Assets',
                'Total',  # Second Total = Total Assets
                'Receivables',
                'Inventory',
                'Cash & Bank',
                'No. of Equity Shares',
                'New Bonus Shares',
                'Face value'
            ]
            
            # Copy Balance Sheet data
            current_target_row = 3
            total_count = 0
            
            for item_name in bs_items:
                bs_ws.cell(current_target_row, 1).value = item_name
                
                # Find this item in source
                for src_row in range(bs_date_row + 1, min(bs_date_row + 25, src_ws.max_row + 1)):
                    src_item = src_ws.cell(src_row, 1).value
                    if src_item:
                        src_item_str = str(src_item).strip()
                        
                        # Handle "Total" - need to track which one
                        if item_name == 'Total':
                            if src_item_str == 'Total':
                                total_count += 1
                                # First Total = row 7, Second Total = row 12
                                if (current_target_row == 7 and total_count == 1) or \
                                   (current_target_row == 12 and total_count == 2):
                                    # Copy data
                                    for idx, src_col in enumerate(date_cols, start=2):
                                        val = src_ws.cell(src_row, src_col).value
                                        bs_ws.cell(current_target_row, idx).value = val
                                    break
                        else:
                            # Normal item matching
                            if src_item_str == item_name:
                                # Copy data
                                for idx, src_col in enumerate(date_cols, start=2):
                                    val = src_ws.cell(src_row, src_col).value
                                    bs_ws.cell(current_target_row, idx).value = val
                                break
                
                current_target_row += 1
            
            # ============== PROFIT AND LOSS ACCOUNT (Second Sheet) ==============
            pl_ws = new_wb.create_sheet("Profit and Loss Account", 1)
            
            # Row 1: Title
            pl_ws['A1'] = 'PROFIT & LOSS'
            
            # Row 2: Report Date header + dates
            pl_ws['A2'] = 'Report Date'
            for idx, date_val in enumerate(dates, start=2):
                pl_ws.cell(2, idx).value = date_val
            
            # Define P&L items IN EXACT ORDER from target
            pl_items = [
                'Sales',
                'Raw Material Cost',
                'Change in Inventory',
                'Power and Fuel',
                'Other Mfr. Exp',
                'Employee Cost',
                'Selling and admin',
                'Other Expenses',
                'Other Income',
                'Depreciation',
                'Interest',
                'Profit before tax',
                'Tax',
                'Net profit',
                'Dividend Amount'
            ]
            
            # Copy P&L data
            current_target_row = 3
            
            for item_name in pl_items:
                pl_ws.cell(current_target_row, 1).value = item_name
                
                # Find this item in source
                for src_row in range(pl_date_row + 1, min(pl_date_row + 30, src_ws.max_row + 1)):
                    src_item = src_ws.cell(src_row, 1).value
                    if src_item and str(src_item).strip() == item_name:
                        # Copy data
                        for idx, src_col in enumerate(date_cols, start=2):
                            val = src_ws.cell(src_row, src_col).value
                            pl_ws.cell(current_target_row, idx).value = val
                        break
                
                current_target_row += 1
            
            # Save
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(screener_excel_path))[0]
                output_path = f"{base_name}_template.xlsx"
            
            new_wb.save(output_path)
            new_wb.close()
            src_wb.close()
            
            print(f"✓ Template created: {output_path}")
            print(f"  - Sheet 1: Balance Sheet ({len(bs_items)} rows)")
            print(f"  - Sheet 2: Profit and Loss Account ({len(pl_items)} rows)")
            
            return output_path
            
        except Exception as e:
            print(f"Error converting: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _setup_template_sheet(self, ws, years, sheet_type):
        """Legacy method - no longer used"""
        pass
    
    def _populate_balance_sheet(self, ws, df, year_row_idx, years):
        """Legacy method - no longer used"""
        pass
    
    def _populate_pl_sheet(self, ws, df, year_row_idx, years):
        """Legacy method - no longer used"""
        pass
    
    def _find_item_values(self, df, item_name, year_row_idx, years):
        """Legacy method - no longer used"""
        pass
    
    def auto_download_and_convert(self, company_symbol, output_dir=".", keep_original=False, use_consolidated=False, use_id_url=False):
        """
        Complete workflow: download, clean, convert to ready-to-use format
        
        Args:
            company_symbol: Company symbol (e.g., 'HONASA') or ID number (e.g., '1285886')
            output_dir: Directory for output files
            keep_original: Keep original downloaded Excel
            use_consolidated: Use consolidated financials
            use_id_url: Use ID-based URL format /company/id/NUMBER/
            
        Returns:
            str: Path to ready-to-use Excel file or None if failed
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # Download Excel
            original_path = os.path.join(output_dir, f"{company_symbol}_original.xlsx")
            downloaded_path = self.download_excel(company_symbol, original_path, use_consolidated, use_id_url)
            
            if not downloaded_path:
                return None
            
            # Remove empty columns
            self.remove_empty_year_columns(downloaded_path)
            
            # Convert to template format
            template_path = os.path.join(output_dir, f"{company_symbol}_template.xlsx")
            converted_path = self.convert_to_template(downloaded_path, template_path)
            
            # Clean up original if not needed
            if not keep_original and os.path.exists(downloaded_path):
                os.remove(downloaded_path)
                print(f"✓ Removed original file")
            
            return converted_path
            
        except Exception as e:
            print(f"Error in workflow: {e}")
            import traceback
            traceback.print_exc()
            return None


def download_screener_data(company_symbol, cookies_path="screener_cookies.pkl", output_dir="."):
    """
    Convenience function for quick downloads
    
    Args:
        company_symbol: Company symbol
        cookies_path: Path to cookies file
        output_dir: Output directory
        
    Returns:
        str: Path to template file
    """
    downloader = ScreenerDownloader(cookies_path)
    return downloader.auto_download_and_convert(company_symbol, output_dir)


if __name__ == "__main__":
    print("Screener Downloader Module")
    print("Usage: from screener_downloader import download_screener_data")
